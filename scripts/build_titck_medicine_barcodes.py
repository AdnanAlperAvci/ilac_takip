import argparse
import json
import re
import sys
import tempfile
import urllib.parse
import urllib.request
from pathlib import Path


TITCK_MEDICINE_LIST_URL = "https://www.titck.gov.tr/dinamikmodul/43"


def normalize_barcode(value):
    digits = "".join(character for character in str(value or "") if character.isdigit())
    if len(digits) == 14 and digits.startswith("0"):
        return digits[1:]

    return digits


def normalize_header(value):
    return str(value or "").strip().casefold()


def normalize_quantity(value):
    text = str(value or "").strip().replace(",", ".")
    match = re.search(r"\d+(?:\.\d+)?", text)
    if match is None:
        return None

    quantity = float(match.group(0))
    if quantity <= 0:
        return None

    return f"{quantity:.2f}"


def extract_box_quantity_from_name(value):
    match = re.search(
        r"(\d+(?:[,.]\d+)?)\s*"
        r"(?:adet|tablet|tbl|kapsül|kapsul|kap|ampul|flakon|saşe|sase|supp|draje|pastil)",
        str(value or ""),
        flags=re.IGNORECASE,
    )
    if match is None:
        return None

    return normalize_quantity(match.group(1))


def find_latest_titck_excel_url(page_url):
    request = urllib.request.Request(
        page_url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36"
            ),
        },
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        html = response.read().decode("utf-8", errors="ignore")

    match = re.search(r'href=["\']([^"\']+\.xlsx?)["\']', html, flags=re.IGNORECASE)
    if match is None:
        raise RuntimeError("TİTCK sayfasında Excel bağlantısı bulunamadı.")

    return urllib.parse.urljoin(page_url, match.group(1))


def download_file(url, destination_path):
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36"
            ),
        },
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        destination_path.write_bytes(response.read())


def find_column(headers, candidates):
    normalized_candidates = {normalize_header(candidate) for candidate in candidates}
    for index, header in enumerate(headers):
        if normalize_header(header) in normalized_candidates:
            return index

    return None


def iter_titck_rows(workbook):
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        header_index = None
        name_column = None
        barcode_column = None
        box_quantity_column = None

        for index, row in enumerate(rows[:10]):
            name_column = find_column(row, ("İlaç Adı", "Ilac Adi", "İlaç adı"))
            barcode_column = find_column(row, ("Barkod", "Barcode"))
            box_quantity_column = find_column(
                row,
                (
                    "Kutu Adedi",
                    "Kutu Miktarı",
                    "Ambalaj Miktarı",
                    "Ambalaj Miktari",
                    "Ambalaj Adedi",
                    "Adet",
                ),
            )
            if name_column is not None and barcode_column is not None:
                header_index = index
                break

        if header_index is None:
            continue

        for row in rows[header_index + 1 :]:
            medicine_name = row[name_column] if name_column < len(row) else None
            barcode = row[barcode_column] if barcode_column < len(row) else None
            box_quantity = (
                row[box_quantity_column]
                if box_quantity_column is not None and box_quantity_column < len(row)
                else None
            )
            yield barcode, medicine_name, box_quantity


def build_barcodes(source_path, output_path):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "Bu script için openpyxl gerekli. Kurulum: python -m pip install openpyxl"
        ) from error

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    barcode_names = {}

    try:
        for barcode, medicine_name, box_quantity in iter_titck_rows(workbook):
            normalized_barcode = normalize_barcode(barcode)
            if normalized_barcode and medicine_name:
                clean_name = str(medicine_name).strip()
                barcode_names[normalized_barcode] = {
                    "name": clean_name,
                    "boxQuantity": normalize_quantity(box_quantity)
                    or extract_box_quantity_from_name(clean_name),
                }
    finally:
        workbook.close()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as output_file:
        json.dump(
            dict(sorted(barcode_names.items())),
            output_file,
            ensure_ascii=False,
            indent=2,
        )
        output_file.write("\n")

    return len(barcode_names)


def parse_args():
    parser = argparse.ArgumentParser(
        description="TİTCK SKRS E-Reçete Excel listesinden barkod asset'i üretir."
    )
    parser.add_argument(
        "--source",
        help="Yerel TİTCK Excel dosyası. Verilmezse resmi TİTCK sayfasından indirilir.",
    )
    parser.add_argument(
        "--output",
        default="assets/medicine_barcodes.json",
        help="Oluşturulacak Flutter asset dosyası.",
    )
    parser.add_argument(
        "--titck-url",
        default=TITCK_MEDICINE_LIST_URL,
        help="TİTCK SKRS E-Reçete liste sayfası.",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    output_path = Path(args.output)

    if args.source:
        source_path = Path(args.source)
        count = build_barcodes(source_path, output_path)
        print(f"{count} barkod kaydı yazıldı: {output_path}")
        return 0

    with tempfile.TemporaryDirectory() as temp_directory:
        source_path = Path(temp_directory) / "titck_medicines.xlsx"
        excel_url = find_latest_titck_excel_url(args.titck_url)
        print(f"TİTCK Excel indiriliyor: {excel_url}")
        download_file(excel_url, source_path)
        count = build_barcodes(source_path, output_path)

    print(f"{count} barkod kaydı yazıldı: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
